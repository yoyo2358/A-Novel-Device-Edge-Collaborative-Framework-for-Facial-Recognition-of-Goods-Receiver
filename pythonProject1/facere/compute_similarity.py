import os
import face_recognition as face
import numpy as np
import openpyxl

frontal_faces = os.listdir('frontal_face200')                        # 加载所有正脸图片（收货人证件照）
frontal_faces_encode_list = []                                      # 定义列表，用来存放正脸的编码
frontal_person_name_list=[]                                         # 用来存放人的身份
for image in frontal_faces:                                         # 循环从列表中读取人脸照片
    frontal_person_name_list.append(image)                          # 将人物身份同步保存
    current_face = face.load_image_file("frontal_face200/" + image)  # 加载要编码的图片
    current_face_encoded = face.face_encodings(current_face)        # 编码
    frontal_faces_encode_list.append(current_face_encoded)          # 挨个儿放入list中
frontal_faces_encode_array=np.array(frontal_faces_encode_list)      # 将list类型变为array类型
print(frontal_faces_encode_array.shape)
frontal_faces_encode_array=np.squeeze(frontal_faces_encode_array)   # 将无用的1个维度压缩掉
print("收货人正面照片加载完成")

name_list1=[]                                                        # 用来存放名字，为了制作Excel表格
frontal_profile_distance_list=[]                                     # 用来存放距离，为了制作Excel表格

profile_faces = os.listdir('profile_face200')                         # 加载侧脸图片（相当于行人图片）
for i1, frontal_person_name in enumerate(frontal_person_name_list):  # 从正脸列表中挨个儿拎儿出行人身份
    print("当前行人身份：" + frontal_person_name +"接下来寻找其对应的人脸编码")
    for i2, frontal_feature in enumerate(frontal_faces_encode_list): # 按照序号去人脸编码列表里找到相应身份的人脸编码
        if i2==i1:                                                   # 如果在侧脸编码列表中找到了
            print("已找到"+frontal_person_name+"的编码,接下来去侧脸身份list去寻找对应的身份")
            flag=0                                                   # 用来判断是否找到想要找的侧脸
            for profile_image in profile_faces:                      # 从侧脸图像中挨个儿拎出来
                if profile_image==frontal_person_name:               #判断当前拎出来的侧脸身份是不是当前对应的正脸身份
                    flag=1
                    pofile_face = face.load_image_file("profile_face200/" + profile_image)  # 加载该侧脸图片
                    profile_faces_encode = face.face_encodings(pofile_face)                # 对其编码
                    frontal_feature_array = np.array(frontal_feature)                      # 将正脸编码数组化
                    distance = face.face_distance(frontal_feature_array, profile_faces_encode) # 计算二者之间的距离
                    print("正脸 "+frontal_person_name+" 和侧脸 "+profile_image+" 之间的距离为："+str(distance))
                    name_list1.append(frontal_person_name)            # 将名字存入
                    frontal_profile_distance_list.append((distance))                 # 将距离存入
            if flag==0:
                print("没有找到对应的人")


name_list2=[]                                                        # 用来存放名字，为了制作Excel表格
frontal_generate_distance_list=[]                                     # 用来存放距离，为了制作Excel表格

generate_faces = os.listdir('generate_face200')                         # 加载侧脸图片（相当于行人图片）
for i1, frontal_person_name in enumerate(frontal_person_name_list):  # 从侧脸列表中挨个儿拎儿出行人身份
    print("当前行人身份：" + frontal_person_name +"接下来寻找其对应的人脸编码")
    for i2, frontal_feature in enumerate(frontal_faces_encode_list): # 按照序号去人脸编码列表里找到相应身份的人脸编码
        if i2==i1:                                                   # 如果在侧脸编码列表中找到了
            print("已找到"+frontal_person_name+"的编码,接下来去侧脸身份list去寻找对应的身份")
            flag=0                                                   # 用来判断是否找到想要找的侧脸
            for generate_image in generate_faces:                      # 从侧脸图像中挨个儿拎出来
                if generate_image==frontal_person_name:               #判断当前拎出来的侧脸身份是不是当前对应的正脸身份
                    flag=1
                    generate_face = face.load_image_file("generate_face200/" + generate_image)  # 加载该侧脸图片
                    generate_faces_encode = face.face_encodings(generate_face)                # 对其编码
                    frontal_feature_array = np.array(frontal_feature)                      # 将正脸编码数组化
                    distance = face.face_distance(frontal_feature_array, generate_faces_encode) # 计算二者之间的距离
                    print("正脸 "+frontal_person_name+" 和转正 "+generate_image+" 之间的距离为："+str(distance))
                    name_list2.append(frontal_person_name)            # 将名字存入
                    frontal_generate_distance_list.append((distance))                 # 将距离存入
            if flag==0:
                print("没有找到对应的人")

# 制作Excel
f = openpyxl.Workbook()
sheet1 = f.create_sheet(index=0)
for i in range(1,len(frontal_profile_distance_list)+1):
    sheet1.cell(i,1,frontal_profile_distance_list[i-1])#书写在第一列
for i in range(1,len(frontal_generate_distance_list)+1):
    sheet1.cell(i,2,frontal_generate_distance_list[i-1])#书写在第二列
f.save('diatance.xlsx')
